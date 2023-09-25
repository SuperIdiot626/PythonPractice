from openpyxl import load_workbook
def main():
    file_location,dire=reference_input()
    excel=load_workbook(file_location,data_only=True)
    table=excel.worksheets[0]
    i=2
    while 1:
        dirname=str(table.cell(row=i, column=5).value)
        if dirname=='None':
            break
        if i==2:
            stringPart1,stringPart2=fileread(dire,dirname)
            if stringPart1=='':
                print('model file error')
                break
        try:
            file=open(dire+'\\'+dirname+'\\mcfd.inp','w')
        except:
            print('dir not found')
            break
        filewriteParts(file,stringPart1)
        file.write('values ')
        file.write(str(table.cell(row=i, column=6).value)+' ')
        file.write(str(table.cell(row=i, column=7).value)+' ')
        file.write(str(table.cell(row=i, column=8).value)+' ')
        file.write(str(table.cell(row=i, column=9).value)+' ')
        file.write(str(table.cell(row=i, column=10).value)+' ')
        file.write('\n')
        filewriteParts(file,stringPart2)
        print('mcfd.inp file written in '+dirname)
        i=i+1
        file.close()
    print(i-2,' files created in corresponding dir')

def fileread(dire,dirname):
    file=open(dire+'\\'+dirname+'\\mcfd.inp','r')
    text=file.readlines()
    for i in text:
        if i[0:7]=='seq.# 2':
            index=text.index(i)
            print(text.index(i))
    stringPart1=(text[:index+1])
    stringPart2=(text[index+2:])
    return stringPart1,stringPart2

def filewriteParts(file,part):
    for i in part:
        file.write(i)

def reference_input():
    file_location=input('please enter the location of your execl file:')
    dire_location=input('please enter the location of your calculation files:')
    file_location+='\\states.xlsx'
    return file_location,dire_location

main()