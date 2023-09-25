from openpyxl import load_workbook
file_location='H:\\0610\\state.xlsx'
excel=load_workbook(file_location)
sheetname='npfoptsinp'
table=excel[sheetname]

stringPart1=r'''options begin
X yes
Y yes
Z yes
Xt no
Yt no
Zt no
P yes
P_Gauge no
Acoustic_Pressure no
R yes
U yes
V yes
W yes
U_Rel no
V_Rel no
W_Rel no
M yes
T yes
Tvib no
Vort_x yes
Vort_y yes
Vort_z yes
P_TOTAL yes
T_TOTAL yes
Enthalpy no
Enthalpy_total no
Energy yes
NumbDens no
MULAM no
MUTUR yes
Cell_distance_function no
Turbulence_index no
Stress_Flatness_Parameter no
GAMMA no
KAPPA no
Volume no
P_x yes
P_y yes
P_z yes
T_x yes
T_y yes
T_z yes
U_x yes
U_y yes
U_z yes
V_x yes
V_y yes
V_z yes
W_x yes
W_y yes
W_z yes
||(R_xyz)|| no
||(R_x)|| no
||(R_y)|| no
||(R_z)|| no
VelMag yes
PreBet no
Strain_rate no
Y_plus yes
Cell_Re yes
Trans_trip no
Cp yes
Heat_release yes
Sp_heat_pres yes
Sp_heat_volu yes
Sound_speed yes
Cp_total no
Qdot yes
HeatTransCoef yes
Nu no
Stanton yes
ThOx_wall_deposit no
Qcrit yes
Kn no
Skin_friction yes
Tau_x yes
Tau_y yes
Tau_z yes
uu_stress yes
vv_stress yes
ww_stress yes
uv_stress yes
uw_stress yes
vw_stress yes
GAMMA_mix no
R_mix no
U_mix no
V_mix no
W_mix no
M_mix no
VolFrac no
Pt_cont no
CPU# yes
Cell_group yes
pinf 79.7791
rinf 1.026407e-03
'''
stringPart2=r'''
edp_eps 1.000000e-08
Nu_tref 2.706500e+02
Nu_lref 1.000000e+00
Kn_lref 1.000000e+00
fv_fsmach 2.000000e-01
fv_alpha 0.000000e+00
fv_reynol 1.000000e+06
Turb1 yes
Turb2 yes
Spec1 yes
MoleFrac1 yes
NumbDens1 yes
NODAVE yes
BC_family yes
cpudomains no
grpdomains no
plt_nomust no
options end

'''
dire='H:\\0610\\0613'

i=1
while(i<=40):
    file=open(dire+'\\'+str(i)+'\\npfopts.inp','w')
    print('file name is '+str(i)+'npfopts.inp')
    file.write(stringPart1)
    file.write('uinf '+str(table.cell(row=i, column=1).value))
    file.write(stringPart2)
    i=i+1
    file.close()
