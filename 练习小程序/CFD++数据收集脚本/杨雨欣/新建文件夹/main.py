from load_zonedat import load_dat
from load_variable import load_vari
import numpy as np
import Up as up
import Under as un
from openpyxl import load_workbook

file_location = r'D:\Three_D_curvature\DAT\design_table_t1.xlsx'
file_location = r'C:\Users\WYZ\Documents\WeChat Files\wxid_jrgz7252a1xb22\FileStorage\MsgAttach\ece701761a9b59c813dce98346280f60\File\2022-06\新建文件夹\design_table_t1(1).xlsx'
excel = load_workbook(file_location)
sheetname = 'DAT'
table = excel[sheetname]

# file_url = input("输入文件路径: ")
file_url=r'D:\Three_D_curvature\DAT\Only-Wall'
file_url=r'C:\Users\WYZ\Documents\WeChat Files\wxid_jrgz7252a1xb22\FileStorage\MsgAttach\ece701761a9b59c813dce98346280f60\File\2022-06\新建文件夹'
L_max = 5000
theta_1 = []
theta_2 = []
Nc1 = []
Nc2 = []
n_value = []
W_max = []
Point_Position_Set = []
Gaussian_Set = []
Mean_Set = []
Normal_Vector_Set = []
Angle_Incoming_Set = []
Dipan_Set = []
Principal_curvature_Set = []
Attack_Angle = 5    #角度制
Slideslip_Angle = 0 #角度制
# f = open("Gaussian.dat", 'w')
for t in range(0,2):
    theta_1.append(table.cell(row=t+1, column=1).value / 180 * np.pi)
    theta_2.append(table.cell(row=t+1, column=2).value / 180 * np.pi)
    W_max.append(table.cell(row=t+1, column=3).value)
    Nc1.append(table.cell(row=t+1, column=4).value)
    Nc2.append(table.cell(row=t+1, column=5).value)
    n_value.append(table.cell(row=t+1, column=6).value)

    variables = load_vari(file_url, t+1)
    zone_data = load_dat(file_url, t+1)
    up_count = 0
    under_count = 0
    zone_data_shape = np.shape(zone_data)
    Point_Position = []
    Gaussian_Curvature = []
    Mean_Curvature = []
    Normal_Vector = []
    Angle_Incoming = []
    Dipan = []
    Principal_curvature = []

    # for i in range(zone_data_shape[0]):
    for i in range(1):
        x = zone_data[i, variables.index('X')]
        y = zone_data[i, variables.index('Y')]
        z = zone_data[i, variables.index('Z')]
        Curvilinear_coordinate = [x, y]
        Point_Position.append(Curvilinear_coordinate)
        if z >= 0:
                up_count += 1
                Gaussian_Curvature.append(up.Gauss_Value(Point_Position[i][0], Point_Position[i][1], theta_1[t], Nc1[t], n_value[t], W_max[t], L_max))
                Mean_Curvature.append(up.Mean_Value(Point_Position[i][0], Point_Position[i][1], theta_1[t], Nc1[t], n_value[t], W_max[t], L_max))
                Normal_Vector.append(up.n_Value(Point_Position[i][0], Point_Position[i][1], theta_1[t], Nc1[t], n_value[t], W_max[t], L_max))
                Angle_Incoming.append(up.Angle_Value(Point_Position[i][0], Point_Position[i][1], theta_1[t], Nc1[t], n_value[t], W_max[t], L_max, Attack_Angle, Slideslip_Angle))
                Dipan.append(up.Dipan_Value(Point_Position[i][0], Point_Position[i][1], theta_1[t], Nc1[t], n_value[t], W_max[t], L_max))
                Principal_curvature.append(up.Principal_curvature_Value(Point_Position[i][0], Point_Position[i][1], theta_1[t], Nc1[t], n_value[t], W_max[t], L_max))
        if z < 0:
                under_count += 1
                Gaussian_Curvature.append(un.Gauss_Value(Point_Position[i][0], Point_Position[i][1], theta_2[t], Nc2[t], n_value[t], W_max[t], L_max))
                Mean_Curvature.append(un.Mean_Value(Point_Position[i][0], Point_Position[i][1], theta_2[t], Nc2[t], n_value[t], W_max[t], L_max))
                Normal_Vector.append(un.n_Value(Point_Position[i][0], Point_Position[i][1], theta_2[t], Nc2[t], n_value[t], W_max[t], L_max))
                Angle_Incoming.append(un.Angle_Value(Point_Position[i][0], Point_Position[i][1], theta_2[t], Nc2[t], n_value[t], W_max[t], L_max, Attack_Angle, Slideslip_Angle))
                Dipan.append(un.Dipan_Value(Point_Position[i][0], Point_Position[i][1], theta_2[t], Nc2[t], n_value[t], W_max[t], L_max))
                Principal_curvature.append(un.Principal_curvature_Value(Point_Position[i][0], Point_Position[i][1], theta_2[t], Nc2[t], n_value[t], W_max[t], L_max))
    Gaussian_Set.append(Gaussian_Curvature)
    Mean_Set.append(Mean_Curvature)
    Normal_Vector_Set.append(Normal_Vector)
    Angle_Incoming_Set.append(Angle_Incoming)
    Dipan_Set.append(Dipan)
    Point_Position_Set.append(Principal_curvature)


    # f.write(Gaussian_Set)
    #
    #
    # print(Gaussian_Set[t])
    # print(Mean_Set[t])
    # print(Angle_Incoming[t])
    # print('Up:', up_count)
    # print('Under:', under_count)

# print(Point_Position)
# print(type(Point_Position))

# print(zone_data[2])
# print(zone_data[2,3])
# print(zone_data[2,variables.index('P')])
# print(Point_Position)
# print(Point_Position[2][0])