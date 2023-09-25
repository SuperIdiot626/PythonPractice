#!/usr/bin/env python3
# -*- coding: utf-8 -*-

#Author Wyz

import openpyxl
import openpyxl.utils


#使用load命令读取excel表格
wb= openpyxl.load_workbook('example.xlsx')  
#print(type(wb))

#a=wb.get_sheet_names()                      #获取shhet名称，返回一个list

a=wb.get_sheet_by_name('example')           #使用sheet名来获取表格


#可通过方括号内部单引号坐标来获取单元格，单元格有value、column、row、coordinate属性
print(a['A1'].value)                        # STATION
print(a['A1'].coordinate)                   # A1
print(a['A1'].column)                       # 1
print(a['A1'].row)                          # 1

#可用两个数字的方式获取单元格，获取的是单元格本身，值与坐标等需另外指定
print(a.cell(1,1).value)

#字母+数字的坐标与数字+数字的坐标可以相互转换，注意，此处使用uyils子库，并非书中的cell
a=openpyxl.utils.get_column_letter(1)               #A
b=openpyxl.utils.get_column_letter(26)              #Z
c=openpyxl.utils.column_index_from_string('ABC')    #731

